"""
Data Preprocessing Pipeline
----------------------------
Reads the raw Excel knowledge base, extracts text from each product sheet,
applies text cleaning, anonymization (PII masking), tokenization-aware
chunking, and outputs clean JSON ready for embedding & indexing.

Usage:
    python -m scripts.preprocessing.preprocess
"""

import json
import os
import re
import uuid

import openpyxl

# ── Configuration ─────────────────────────────────────────────────────
INPUT_PATH = os.path.join("NUST Bank-Product-Knowledge.xlsx")
OUTPUT_DIR = os.path.join("data")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "preprocessed_chunks.json")

SKIP_SHEETS = {"Main", "Rate Sheet July 1 2024", "Sheet1"}

# ── PII patterns for anonymization ───────────────────────────────────
PII_PATTERNS = [
    (re.compile(r"\b\d{5}[-\s]?\d{7}[-\s]?\d{1}\b"), "[CNIC_REDACTED]"),
    (re.compile(r"\b\d{13}\b"), "[ID_REDACTED]"),
    (re.compile(r"\b\d{16}\b"), "[CARD_REDACTED]"),
    (re.compile(r"\b\d{4}[-\s]\d{4}[-\s]\d{4}[-\s]\d{4}\b"), "[CARD_REDACTED]"),
    (re.compile(r"\b[A-Z]{2}\d{2}[A-Z0-9]{11,30}\b"), "[IBAN_REDACTED]"),
    (re.compile(r"\b[\w.-]+@[\w.-]+\.\w{2,}\b"), "[EMAIL_REDACTED]"),
    (re.compile(r"\b(?:\+92|0)\d{3}[-\s]?\d{7}\b"), "[PHONE_REDACTED]"),
]


def anonymize(text: str) -> str:
    """Mask any PII found in the text."""
    for pattern, replacement in PII_PATTERNS:
        text = pattern.sub(replacement, text)
    return text


def clean_text(text: str) -> str:
    """Normalize whitespace, strip control characters, lowercase."""
    if not text:
        return ""
    # Remove non-breaking spaces and other control characters
    text = text.replace("\xa0", " ")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    # Collapse multiple spaces / newlines
    text = re.sub(r"\s+", " ", text).strip()
    # Lowercase
    text = text.lower()
    return text


def extract_rows_from_sheet(ws) -> list[str]:
    """
    Extract meaningful text rows from a worksheet.
    Reads columns A and B (where the content lives), skips empty rows
    and formula references (strings starting with '=').
    """
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Gather non-empty cell values from first two columns
        cells = []
        for cell in row[:2]:  # columns A and B
            if cell is None:
                continue
            val = str(cell).strip()
            if not val or val.startswith("="):
                continue
            cells.append(val)
        if cells:
            rows.append(" ".join(cells))
    return rows


def chunk_sheet(sheet_name: str, rows: list[str]) -> list[dict]:
    """
    Convert extracted rows into chunks.
    Groups Q&A pairs (question row followed by answer row) and
    standalone facts into individual chunks.
    """
    chunks = []
    i = 0
    while i < len(rows):
        text = rows[i]
        cleaned = clean_text(text)

        # Skip very short or header-only rows
        if len(cleaned) < 5:
            i += 1
            continue

        # If current row looks like a question, pair it with the next row
        if cleaned.endswith("?") and i + 1 < len(rows):
            answer = clean_text(rows[i + 1])
            if answer and len(answer) >= 5:
                combined = f"q: {cleaned} a: {answer}"
                combined = anonymize(combined)
                chunks.append({
                    "id": str(uuid.uuid4()),
                    "account": sheet_name,
                    "text": combined,
                })
                i += 2
                continue

        # Standalone fact / row
        cleaned = anonymize(cleaned)
        chunks.append({
            "id": str(uuid.uuid4()),
            "account": sheet_name,
            "text": cleaned,
        })
        i += 1

    return chunks


def preprocess():
    """Main preprocessing pipeline."""
    print(f"Loading workbook: {INPUT_PATH}")
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)

    product_sheets = [s for s in wb.sheetnames if s not in SKIP_SHEETS]
    print(f"Found {len(product_sheets)} product sheets to process")

    all_chunks = []

    for sheet_name in product_sheets:
        ws = wb[sheet_name]
        rows = extract_rows_from_sheet(ws)
        chunks = chunk_sheet(sheet_name, rows)
        print(f"  {sheet_name:20s} → {len(rows):3d} rows → {len(chunks):3d} chunks")
        all_chunks.extend(chunks)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_chunks, f, indent=2, ensure_ascii=False)

    print(f"\nTotal chunks: {len(all_chunks)}")
    print(f"Output saved to: {OUTPUT_PATH}")
    return all_chunks


if __name__ == "__main__":
    preprocess()
